import os
import stat
import paramiko
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import time
import pickle
import subprocess

def listar_archivos_remotos_linux(sftp_client, carpeta):
    archivos = []
    try:
        for entry in sftp_client.listdir_attr(carpeta):
            ruta_completa = carpeta + "/" + entry.filename
            if stat.S_ISDIR(entry.st_mode):
                archivos.extend(listar_archivos_remotos_linux(sftp_client, ruta_completa))
            else:
                fecha_modificacion = datetime.fromtimestamp(entry.st_mtime)
                fecha_acceso = datetime.fromtimestamp(entry.st_atime)
                nombre, extension = os.path.splitext(entry.filename)
                ruta_padre = os.path.dirname(ruta_completa)
                archivos.append((nombre, extension, fecha_modificacion, fecha_acceso, ruta_padre))
    except FileNotFoundError as e:
        print(f"No se encontró la carpeta remota: {carpeta} - {e}")
    except PermissionError as e:
        print(f"Error de permisos con la carpeta: {carpeta} - {e}")
    except Exception as e:
        print(f"Error inesperado al listar archivos remotos: {e}")
    return archivos

def listar_archivos_remotos_windows(sftp_client, carpeta):
    archivos = []
    try:
        for entry in sftp_client.listdir_attr(carpeta):
            ruta_completa = os.path.join(carpeta, entry.filename.replace('/', '\\'))
            if stat.S_ISDIR(entry.st_mode):
                archivos.extend(listar_archivos_remotos_windows(sftp_client, ruta_completa))
            else:
                fecha_modificacion = datetime.fromtimestamp(entry.st_mtime)
                fecha_acceso = datetime.fromtimestamp(entry.st_atime)
                nombre, extension = os.path.splitext(entry.filename)
                ruta_padre = os.path.dirname(ruta_completa)
                archivos.append((nombre, extension, fecha_modificacion, fecha_acceso, ruta_padre))
    except FileNotFoundError as e:
        print(f"No se encontró la carpeta remota: {carpeta} - {e}")
    except PermissionError as e:
        print(f"Error de permisos con la carpeta: {carpeta} - {e}")
    except Exception as e:
        print(f"Error inesperado al listar archivos remotos: {e}")
    return archivos

def listar_archivos_remotos_windows_red_compartida(ruta_red):
    """Obtiene las rutas compartidas y lista archivos en ellas, ignorando las carpetas locales con '$' en su nombre."""
    archivos_totales = []
    try:
        # Obtener rutas compartidas
        resultado = subprocess.check_output(["powershell", "-Command", "Get-SmbShare"], text=True)
        lineas = resultado.strip().split('\n')
        
        # Extraer los nombres de las rutas compartidas
        rutas_compartidas = []
        for linea in lineas[3:]:  # Saltar las primeras 3 líneas de encabezado
            columnas = linea.split()
            if len(columnas) > 0:
                nombre = columnas[0]
                if '$' not in nombre:
                    rutas_compartidas.append(columnas[2])  # Tomar el valor de la columna 'Path'

        for ruta_red in rutas_compartidas:
            ruta_red = ruta_red.strip()  # Limpiar espacios en blanco

            if not ruta_red:
                continue

            # Validar que la ruta compartida sea accesible
            if os.path.exists(ruta_red):
                # Verificar si la ruta es accesible
                ruta_red = os.path.normpath(ruta_red)
                for dirpath, dirnames, filenames in os.walk(ruta_red):
                    for nombre_archivo in filenames:
                        ruta_completa = os.path.join(dirpath, nombre_archivo)
                        fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(ruta_completa))
                        fecha_acceso = datetime.fromtimestamp(os.path.getatime(ruta_completa))
                        nombre, extension = os.path.splitext(nombre_archivo)
                        ruta_padre = os.path.dirname(ruta_completa)
                        archivos_totales.append((nombre, extension, fecha_modificacion, fecha_acceso, ruta_padre))
            else:
                print(f"No se puede acceder a la ruta compartida: {ruta_red}")

    except subprocess.CalledProcessError as e:
        print(f"Error al obtener las rutas compartidas: {e}")
    except Exception as e:
        print(f"Error al listar archivos en la red compartida: {e}")

    return archivos_totales

def agregar_hoja_excel(libro_excel, nombre_hoja):
    nombre_hoja = limpiar_nombre(nombre_hoja)
    if nombre_hoja in libro_excel.sheetnames:
        print(f"La hoja '{nombre_hoja}' ya existe en el archivo de Excel. Es necesario cambiar el nombre.")
        nuevo_nombre = input("Ingrese un nuevo nombre para la hoja: ")
        return agregar_hoja_excel(libro_excel, nuevo_nombre)
    else:
        libro_excel.create_sheet(title=nombre_hoja)
        return nombre_hoja

def ajustar_ancho_columnas(hoja):
    for col in hoja.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja.column_dimensions[column].width = adjusted_width

def guardar_en_excel(archivos, nombre_archivo, nombre_hoja_nueva=None):
    # Determina la ruta del archivo en la misma carpeta donde se ejecuta el código
    nombre_archivo = nombre_archivo or "Bitacora.xlsx"
    nombre_archivo = nombre_archivo if nombre_archivo.lower().endswith('.xlsx') else nombre_archivo + '.xlsx'
    ruta_archivo = os.path.join(os.getcwd(), nombre_archivo)

    if os.path.exists(ruta_archivo):
        libro_excel = load_workbook(ruta_archivo)
    else:
        libro_excel = Workbook()

    nombre_hoja = nombre_hoja_nueva if nombre_hoja_nueva else "Datos"
    nombre_hoja_final = agregar_hoja_excel(libro_excel, nombre_hoja)
    hoja = libro_excel[nombre_hoja_final]
    
    # Añadir cabecera si es una nueva hoja
    if hoja.max_row == 1:
        hoja.append(["ID", "Nombre de dato / archivo", "Formato", "Fecha de creación", "Fecha de modificación", "Ruta",
                     "Responsable del dato / archivo", "Propósito del dato / archivo",
                     "¿Quién tiene acceso al archivo / dato?", "Transferencia con 3ero / externos",
                     "Responsable de respaldo", "Fecha de respaldo", "Responsable de eliminación", "Fecha de eliminación"])

    # Iniciar cronómetro
    tiempo_inicio = time.time()

    for id_archivo, archivo in enumerate(archivos, start=1):
        hoja.append([id_archivo] + list(archivo))
    
    ajustar_ancho_columnas(hoja)
    
    try:
        libro_excel.save(ruta_archivo)
        tiempo_total = time.time() - tiempo_inicio
        print(f"Datos guardados exitosamente en el archivo Excel: {ruta_archivo}")
        print(f"Tiempo total para guardar en Excel: {tiempo_total:.2f} segundos")
    except PermissionError as e:
        print(f"Error de permisos al guardar el archivo: {ruta_archivo} - {e}")

def limpiar_nombre(nombre):
    caracteres_no_validos = ['\\', '/', '*', '[', ']', ':', '?']
    for caracter in caracteres_no_validos:
        nombre = nombre.replace(caracter, '_')
    return nombre[:30]

def establecer_conexion(host, username, use_private_key, private_key_path=None, password=None, port=22, passphrase=None):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    
    try:
        if use_private_key and private_key_path:
            pkey = paramiko.RSAKey.from_private_key_file(private_key_path, password=passphrase)
            ssh.connect(hostname=host, port=port, username=username, pkey=pkey)
        else:
            ssh.connect(hostname=host, port=port, username=username, password=password)
        print(f"Conexión establecida con {host}")
        return ssh
    except paramiko.AuthenticationException:
        print("Fallo en la autenticación")
    except paramiko.SSHException as sshException:
        print(f"Error al intentar conectar al host: {sshException}")
    except Exception as e:
        print(f"Error inesperado al intentar conectar: {e}")
        return None

def guardar_progreso(progreso, nombre_archivo="progreso.pkl"):
    with open(nombre_archivo, 'wb') as archivo:
        pickle.dump(progreso, archivo)
    print(f"Progreso guardado en {nombre_archivo}.")

def cargar_progreso(nombre_archivo="progreso.pkl"):
    if os.path.exists(nombre_archivo):
        with open(nombre_archivo, 'rb') as archivo:
            return pickle.load(archivo)
    return None

def iniciar_operacion_multiple(conexiones, nombre_archivo_excel, hora_final=None, continuar_progreso=False):
    progreso = cargar_progreso() if continuar_progreso else {}

    # Convertir hora_final en un objeto datetime solo si se proporciona
    if hora_final:
        hora_final_dt = datetime.combine(datetime.today(), hora_final)
    else:
        hora_final_dt = None

    # Establecer hora_fin como una fecha y hora completa (por defecto 5 AM del siguiente día)
    hora_fin = datetime.now().replace(hour=5, minute=0, second=0, microsecond=0)
    if datetime.now() >= hora_fin:
        hora_fin += timedelta(days=1)

    for conexion in conexiones:
        host, username, use_private_key, private_key_path, password, port, rutas, sistema_operativo, passphrase = conexion
        ssh_client = establecer_conexion(host, username, use_private_key, private_key_path, password, port, passphrase)
        if ssh_client:
            try:
                sftp_client = ssh_client.open_sftp()
                for carpeta, nombre_hoja_nueva, es_red_compartida in rutas:
                    progreso_clave = f"{host}_{carpeta}"
                    if progreso_clave in progreso:
                        print(f"Saltando carpeta {carpeta} en {host} ya que se completó en una ejecución previa.")
                        continue

                    ahora = datetime.now()

                    # Comprobar si se ha alcanzado la hora límite
                    if ahora >= hora_fin:
                        print("Se alcanzó la hora límite. Pausando la operación.")
                        guardar_progreso(progreso)
                        return
                    
                    # Comprobar si se ha alcanzado la hora final (solo si se ha proporcionado)
                    if hora_final_dt and ahora >= hora_final_dt:
                        print("Se alcanzó la hora final. Pausando la operación.")
                        guardar_progreso(progreso)
                        return
                    
                    if es_red_compartida:
                        if sistema_operativo == "windows":
                            archivos = listar_archivos_remotos_windows_red_compartida(carpeta)
                        else:
                            print("Las carpetas de red compartidas no son compatibles con sistemas operativos que no sean Windows.")
                            continue
                    else:
                        if sistema_operativo == "linux":
                            archivos = listar_archivos_remotos_linux(sftp_client, carpeta)
                        elif sistema_operativo == "windows":
                            archivos = listar_archivos_remotos_windows(sftp_client, carpeta)
                        else:
                            print("Sistema operativo no soportado.")
                            continue
                    
                    if archivos:
                        guardar_en_excel(archivos, nombre_archivo_excel, nombre_hoja_nueva)
                        print(f"Datos guardados exitosamente en la hoja '{nombre_hoja_nueva}' en {nombre_archivo_excel}.")
                    else:
                        print(f"No se encontraron archivos en la carpeta {carpeta}.")
                    progreso[progreso_clave] = True
            except Exception as e:
                print(f"Error al procesar la operación en {host}: {e}")
            finally:
                ssh_client.close()
        else:
            print(f"No se pudo establecer la conexión SSH con {host}.")

    guardar_progreso(progreso)

def esperar_hasta_hora_objetivo(hora_objetivo):
    ahora = datetime.now()
    objetivo = ahora.replace(hour=hora_objetivo.hour, minute=hora_objetivo.minute, second=0, microsecond=0)
    
    if ahora > objetivo:
        objetivo = objetivo + timedelta(days=1)
    
    tiempo_espera = (objetivo - ahora).total_seconds()
    print(f"Esperando {tiempo_espera / 60:.2f} minutos para ejecutar la tarea a las {objetivo.time()}.")
    time.sleep(tiempo_espera)

def main():
    def obtener_entrada_opcion(prompt, opciones_validas):
        """Obtiene una opción válida del usuario."""
        while True:
            entrada = input(prompt).strip().lower()
            if entrada in opciones_validas:
                return entrada
            else:
                print(f"Opción no válida. Por favor, ingrese {' o '.join(opciones_validas)}.")

    def obtener_entrada_numerica(prompt, valor_default):
        """Obtiene una entrada numérica válida del usuario."""
        while True:
            entrada = input(prompt).strip()
            if entrada == "":
                return valor_default
            if entrada.isdigit():
                return int(entrada)
            else:
                print("Entrada no válida. Por favor, ingrese un número.")

    def obtener_entrada_no_vacia(prompt):
        """Obtiene una entrada no vacía del usuario."""
        while True:
            entrada = input(prompt).strip()
            if entrada:
                return entrada
            else:
                print("La entrada no puede estar vacía.")

    def obtener_entrada_hora(prompt):
        """Obtiene una hora en formato HH:MM del usuario y valida el formato."""
        while True:
            entrada = input(prompt).strip()
            try:
                return datetime.strptime(entrada, "%H:%M").time()
            except ValueError:
                print("El formato de la hora no es válido. Asegúrese de usar HH:MM.")

    conexiones = []
    while True:
        host = obtener_entrada_no_vacia("Ingrese la dirección IP o hostname del servidor: ")
        username = obtener_entrada_no_vacia("Ingrese el nombre de usuario para la conexión SSH: ")
        use_private_key = obtener_entrada_opcion("¿Desea usar una clave privada para la autenticación? (s/n): ", ['s', 'n']) == 's'
        private_key_path = input("Ingrese la ruta al archivo de clave privada (deje vacío si no usa clave privada): ") or None
        passphrase = input("Ingrese la passphrase para la clave privada (deje vacío si no requiere passphrase): ") or None
        password = input("Ingrese la contraseña (en caso de necesitarla, en caso contrario dejar vacio): ") or None
        port = obtener_entrada_numerica("Ingrese el puerto del servidor SSH (por defecto 22): ", 22)
        
        while True:
            sistema_operativo = obtener_entrada_opcion("Ingrese el sistema operativo (linux/windows): ", ["linux", "windows"])
            if sistema_operativo in ["linux", "windows"]:
                break

        rutas = []
        nombres_hojas = set()  # Usamos un conjunto para verificar unicidad
        while True:
            carpeta = obtener_entrada_no_vacia("Ingrese la carpeta remota a analizar: ")
            nombre_hoja_nueva = obtener_entrada_no_vacia("Ingrese el nombre de la nueva hoja: ")
            
            # Verificar si el nombre de la hoja ya está en uso
            while nombre_hoja_nueva in nombres_hojas:
                print(f"La hoja '{nombre_hoja_nueva}' ya existe. Por favor, ingrese un nombre diferente.")
                nombre_hoja_nueva = obtener_entrada_no_vacia("Ingrese un nuevo nombre para la hoja: ")
            
            nombres_hojas.add(nombre_hoja_nueva)
            es_red_compartida = obtener_entrada_opcion("¿Es esta una carpeta de red compartida? (s/n): ", ['s', 'n']) == 's'
            rutas.append((carpeta, nombre_hoja_nueva, es_red_compartida))
            otra_ruta = obtener_entrada_opcion("¿Desea agregar otra ruta a analizar para este servidor? (s/n): ", ['s', 'n'])
            if otra_ruta != 's':
                break
        
        conexiones.append((host, username, use_private_key, private_key_path, password, port, rutas, sistema_operativo, passphrase))
        otro_servidor = obtener_entrada_opcion("¿Desea agregar otro servidor a analizar? (s/n): ", ['s', 'n'])
        if otro_servidor != 's':
            break

    nombre_archivo_excel = input("Ingrese el nombre del archivo Excel (deje vacío para 'Bitacora.xlsx'): ") or "Bitacora.xlsx"
    continuar_progreso = obtener_entrada_opcion("¿Desea continuar desde un progreso guardado? (s/n): ", ['s', 'n']) == 's'
    opcion = obtener_entrada_opcion("¿Desea ejecutar la operación ahora (1) o en una hora específica (2)? Ingrese 1 o 2: ", ['1', '2'])
    
    if opcion == '1':
        # Ejecutar la operación de inmediato sin establecer un tiempo límite
        iniciar_operacion_multiple(conexiones, nombre_archivo_excel, hora_final=None, continuar_progreso=continuar_progreso)
    
    elif opcion == '2':
        hora_inicio = obtener_entrada_hora("Ingrese la hora de inicio en formato HH:MM (por ejemplo, 22:00): ")
        print(f"Hora de inicio establecida a las {hora_inicio}.")
        
        hora_final_str = obtener_entrada_opcion("¿Desea agregar una hora final para limitar el tiempo de ejecución? (s/n): ", ['s', 'n'])
        if hora_final_str == 's':
            hora_final = obtener_entrada_hora("Ingrese la hora final en formato HH:MM (por ejemplo, 05:00): ")
            print(f"Hora final establecida a las {hora_final}.")
        else:
            hora_final = None

        esperar_hasta_hora_objetivo(hora_inicio)
        iniciar_operacion_multiple(conexiones, nombre_archivo_excel, hora_final=hora_final, continuar_progreso=continuar_progreso)
    
    else:
        print("Opción no válida. Por favor, ingrese 1 o 2.")

if __name__ == "__main__":
    print(
'''
_____________________$$$
____________________$___$               
_____________________$$$
_____________________$_$
_____________________$_$
___________________$$$_$$$
_________________$$__$$$__$$$
_______________$$__$$$$$$$___$
______________$_______________$
_____________$_________________$
_____________$_________________$
_____________$_____$$$$$$$$$$$$$$$
_____________$____$_______________$
_____________$____$___$$$$$$$$$$$$$
_____________$___$___$___________$$$
_____________$___$___$_$$$___$$$__$$
_____________$___$___$_$$$___$$$__$$
_____________$___$___$___________$$$
_____________$____$___$$$$$$$$$$$$$
_____________$_____$$$$$$$$$$$$$$
_____________$_________________$
_____________$____$$$$$$$$$$$$$$
_____________$___$__$__$__$__$
_____________$__$$$$$$$$$$$$$$
_____________$__$___$__$__$__$
_____________$___$$$$$$$$$$$$$$$
____________$$$_________________$$$
__________$$___$$$_________$$$$$___$$
________$$________$$$$$$$$$__________$$$
_______$__$$_____________________$$$$___$$
____$$$$$___$$$$$$$$______$$$$$$$_______$_$
__$______$$_________$$$$$$______________$_$$
_$____$____$____________________________$_$_$
_$_____$___$______________$$$$$$$$$$$___$_$_$$
_$$$____$___$__$$$$$$$$$$$$__________$___$_$_$$
$___$$$$____$__$_____________________$___$_$$_$
$$$____$___$$__$_____________________$$__$_$__$
$___$__$__$$___$______________________$__$$$__$
$_____$$_$$____$_______________$$$____$__$_$__$

--- Proceso Automático Unificado para Listado e Inventariado de Nodos Activos ---

''')
    main()
