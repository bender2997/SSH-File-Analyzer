import os
import stat
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import time

def listar_archivos_local(carpeta):
    """
    Lista los archivos en una carpeta local.
    
    :param carpeta: Ruta de la carpeta local a listar.
    :return: Lista de tuplas con información de archivos (nombre, extensión, fecha de modificación, fecha de acceso, ruta del padre).
    """
    archivos = []
    try:
        for root, dirs, files in os.walk(carpeta):
            for nombre_archivo in files:
                ruta_completa = os.path.join(root, nombre_archivo)
                fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(ruta_completa))
                fecha_acceso = datetime.fromtimestamp(os.path.getatime(ruta_completa))
                nombre, extension = os.path.splitext(nombre_archivo)
                ruta_padre = os.path.dirname(ruta_completa)
                archivos.append((nombre, extension, fecha_modificacion, fecha_acceso, ruta_padre))
    except FileNotFoundError as e:
        print(f"No se encontró la carpeta local: {carpeta} - {e}")
    except PermissionError as e:
        print(f"Error de permisos con la carpeta local: {carpeta} - {e}")
    except Exception as e:
        print(f"Error inesperado al listar archivos locales: {e}")
    return archivos

def agregar_hoja_excel(libro_excel, nombre_hoja):
    """
    Agrega una nueva hoja a un libro de Excel si el nombre de la hoja no está ya en uso.
    Si el nombre ya existe, solicita al usuario un nuevo nombre.

    :param libro_excel: Objeto Workbook de openpyxl.
    :param nombre_hoja: Nombre deseado para la nueva hoja.
    :return: Nombre de la hoja que fue creada.
    """
    nombre_hoja = limpiar_nombre(nombre_hoja)
    if nombre_hoja in libro_excel.sheetnames:
        print(f"La hoja '{nombre_hoja}' ya existe en el archivo de Excel. Es necesario cambiar el nombre.")
        nuevo_nombre = input("Ingrese un nuevo nombre para la hoja: ")
        return agregar_hoja_excel(libro_excel, nuevo_nombre)
    else:
        libro_excel.create_sheet(title=nombre_hoja)
        return nombre_hoja

def ajustar_ancho_columnas(hoja):
    """
    Ajusta el ancho de las columnas de una hoja de Excel para que el contenido se ajuste automáticamente.

    :param hoja: Objeto Worksheet de openpyxl.
    """
    for col in hoja.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja.column_dimensions[column].width = adjusted_width

def guardar_en_excel(archivos, nombre_archivo, nombre_hoja_nueva=None):
    """
    Guarda la lista de archivos en un archivo de Excel. Crea un nuevo archivo si no existe.
    También crea una nueva hoja si se proporciona un nombre para ella.

    :param archivos: Lista de tuplas con información de archivos.
    :param nombre_archivo: Nombre del archivo Excel donde se guardarán los datos.
    :param nombre_hoja_nueva: Nombre de la hoja donde se guardarán los datos (opcional).
    """
    if not nombre_archivo:
        nombre_archivo = "Bitacora.xlsx"
    if not nombre_archivo.lower().endswith('.xlsx'):
        nombre_archivo += '.xlsx'
    if os.path.exists(nombre_archivo):
        libro_excel = load_workbook(nombre_archivo)
    else:
        libro_excel = Workbook()
    nombre_hoja = nombre_hoja_nueva if nombre_hoja_nueva else "Datos"
    nombre_hoja_final = agregar_hoja_excel(libro_excel, nombre_hoja)
    hoja = libro_excel[nombre_hoja_final]
    hoja.append(["ID", "Nombre de dato / archivo", "Formato", "Fecha de creación", "Fecha de modificación", "Ruta",
                 "Responsable del dato / archivo", "Propósito del dato / archivo",
                 "¿Quién tiene acceso al archivo / dato?", "Transferencia con 3ero / externos",
                 "Responsable de respaldo", "Fecha de respaldo", "Responsable de eliminación", "Fecha de eliminación"])
    for id_archivo, archivo in enumerate(archivos, start=1):
        hoja.append([id_archivo] + list(archivo))
    ajustar_ancho_columnas(hoja)
    try:
        libro_excel.save(nombre_archivo)
        print("Datos guardados exitosamente en el archivo de Excel.")
    except PermissionError as e:
        print(f"Error de permisos al guardar el archivo: {nombre_archivo} - {e}")

def limpiar_nombre(nombre):
    """
    Limpia el nombre de una hoja de Excel eliminando caracteres no válidos.

    :param nombre: Nombre original de la hoja.
    :return: Nombre de la hoja limpio y truncado a 30 caracteres.
    """
    caracteres_no_validos = ['\\', '/', '*', '[', ']', ':', '?']
    for caracter in caracteres_no_validos:
        nombre = nombre.replace(caracter, '_')
    return nombre[:30]

def iniciar_operacion():
    """
    Función principal que lista archivos locales y guarda la información en un archivo de Excel.
    Incluye un contador de tiempo para medir la duración de la ejecución.
    """
    start_time = time.time()  # Inicia el contador de tiempo

    carpeta = input("Ingrese la ruta de la carpeta local a analizar: ")
    nombre_archivo_excel = input("Ingrese el nombre del archivo Excel (deje vacío para 'Bitacora.xlsx'): ") or "Bitacora.xlsx"
    nombre_hoja_nueva = input("Ingrese el nombre de la nueva hoja: ")

    archivos = listar_archivos_local(carpeta)
    guardar_en_excel(archivos, nombre_archivo_excel, nombre_hoja_nueva)

    end_time = time.time()  # Finaliza el contador de tiempo
    elapsed_time = end_time - start_time
    print(f"Tiempo total de ejecución: {elapsed_time:.2f} segundos")

if __name__ == "__main__":
    iniciar_operacion()
