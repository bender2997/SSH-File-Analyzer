import os
import stat
import paramiko
import time
import pickle
import subprocess
import threading
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def listar_archivos_remotos_linux(sftp_client, carpeta, detener_event):
    archivos = []
    try:
        for entry in sftp_client.listdir_attr(carpeta):
            # Verificación del evento para detener el proceso
            if detener_event.is_set():
                print("Proceso detenido por el evento de detención.")
                return archivos

            ruta_completa = carpeta + "/" + entry.filename
            if stat.S_ISDIR(entry.st_mode):
                archivos.extend(listar_archivos_remotos_linux(sftp_client, ruta_completa, detener_event))
            else:
                fecha_modificacion = datetime.fromtimestamp(entry.st_mtime)
                fecha_acceso = datetime.fromtimestamp(entry.st_atime)
                nombre, extension = os.path.splitext(entry.filename)
                ruta_padre = os.path.dirname(ruta_completa)
                archivos.append((nombre, extension, fecha_modificacion, fecha_acceso, ruta_padre))
    except FileNotFoundError as e:
        print(f"No se encontró la carpeta remota: {carpeta} - {e}")
    except PermissionError as e:
        print(f"Error de permisos con la carpeta: {carpeta} - {e}")
    except Exception as e:
        print(f"Error inesperado al listar archivos remotos: {e}")
    return archivos

def listar_archivos_remotos_windows(sftp_client, carpeta, detener_event):
    archivos = []
    try:
        for entry in sftp_client.listdir_attr(carpeta):
            # Verificación del evento para detener el proceso
            if detener_event.is_set():
                print("Proceso detenido por el evento de detención.")
                return archivos

            ruta_completa = os.path.join(carpeta, entry.filename.replace('/', '\\'))
            if stat.S_ISDIR(entry.st_mode):
                archivos.extend(listar_archivos_remotos_windows(sftp_client, ruta_completa, detener_event))
            else:
                fecha_modificacion = datetime.fromtimestamp(entry.st_mtime)
                fecha_acceso = datetime.fromtimestamp(entry.st_atime)
                nombre, extension = os.path.splitext(entry.filename)
                ruta_padre = os.path.dirname(ruta_completa)
                archivos.append((nombre, extension, fecha_modificacion, fecha_acceso, ruta_padre))
    except FileNotFoundError as e:
        print(f"No se encontró la carpeta remota: {carpeta} - {e}")
    except PermissionError as e:
        print(f"Error de permisos con la carpeta: {carpeta} - {e}")
    except Exception as e:
        print(f"Error inesperado al listar archivos remotos: {e}")
    return archivos

def listar_archivos_remotos_windows_red_compartida(ruta_red, detener_event):
    """Obtiene las rutas compartidas y lista archivos en ellas, ignorando las carpetas locales con '$' en su nombre."""
    archivos_totales = []
    try:
        # Obtener rutas compartidas desde el servidor
        resultado = subprocess.check_output(["powershell", "-Command", "Get-SmbShare"], text=True)
        lineas = resultado.strip().split('\n')

        # Extraer los nombres y rutas de las carpetas compartidas
        rutas_compartidas = []
        for linea in lineas[3:]:  # Saltar las primeras 3 líneas de encabezado
            columnas = linea.split()
            if len(columnas) > 0:
                nombre = columnas[0]
                if '$' not in nombre:  # Ignorar carpetas administrativas
                    # Concatenar la ruta de red con la carpeta compartida
                    rutas_compartidas.append(f"{ruta_red}\\{nombre}")

        # Si se ingresa una ruta específica (ruta_red), usar solo esa para la búsqueda
        if os.path.exists(ruta_red):
            rutas_compartidas = [ruta_red]  # Sobrescribir para analizar solo la ruta ingresada

        # Analizar cada carpeta compartida accesible
        for ruta_compartida in rutas_compartidas:
            ruta_compartida = ruta_compartida.strip()  # Limpiar espacios en blanco

            # Verificación del evento para detener el proceso
            if detener_event.is_set():
                print("Proceso detenido por el evento de detención.")
                return archivos_totales

            if not ruta_compartida:
                continue

            # Validar que la ruta compartida sea accesible
            if os.path.exists(ruta_compartida):
                # Convertir la ruta en un formato legible para el sistema operativo
                ruta_compartida = os.path.normpath(ruta_compartida)

                # Recorrer archivos en la ruta compartida
                for dirpath, dirnames, filenames in os.walk(ruta_compartida):
                    for nombre_archivo in filenames:
                        # Verificación del evento para detener el proceso
                        if detener_event.is_set():
                            print("Proceso detenido por el evento de detención.")
                            return archivos_totales

                        # Generar la ruta completa en formato de red
                        ruta_completa_servidor = os.path.join(ruta_red, os.path.relpath(dirpath, ruta_compartida), nombre_archivo)
                        
                        # Obtener las fechas de modificación y acceso
                        ruta_completa_local = os.path.join(dirpath, nombre_archivo)
                        fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(ruta_completa_local))
                        fecha_acceso = datetime.fromtimestamp(os.path.getatime(ruta_completa_local))
                        nombre, extension = os.path.splitext(nombre_archivo)
                        ruta_padre = os.path.dirname(ruta_completa_local)

                        # Agregar a la lista con la ruta completa del servidor
                        archivos_totales.append((nombre, extension, fecha_modificacion, fecha_acceso, ruta_padre))
            else:
                print(f"No se puede acceder a la ruta compartida: {ruta_compartida}")

    except subprocess.CalledProcessError as e:
        print(f"Error al obtener las rutas compartidas: {e}")
    except Exception as e:
        print(f"Error al listar archivos en la red compartida: {e}")

    return archivos_totales

def agregar_hoja_excel(libro_excel, nombre_hoja):
    # Verificar si se ha dado un nombre de hoja válido
    if not nombre_hoja:
        return None  # Si el nombre está vacío, no hacer nada

    # Verificar si la hoja ya existe
    if nombre_hoja in libro_excel.sheetnames:
        return nombre_hoja  # No crear una nueva, solo devolver el nombre de la hoja existente
    
    libro_excel.create_sheet(title=nombre_hoja)
    return nombre_hoja

def ajustar_ancho_columnas(hoja):
    for col in hoja.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja.column_dimensions[column].width = adjusted_width

# Lista global para almacenar el resumen del análisis de cada servidor
resumen_analisis_servidores = []

def guardar_en_excel(archivos, nombre_archivo, ip_servidor, ruta_servidor, nombre_hoja_nueva=None):
    # Determina la ruta del archivo en la misma carpeta donde se ejecuta el código
    nombre_archivo = nombre_archivo or "Bitacora.xlsx"
    ruta_archivo = os.path.join(os.getcwd(), nombre_archivo)

    # Si el archivo ya existe, abrirlo; si no, crear uno nuevo
    if os.path.exists(ruta_archivo):
        libro_excel = load_workbook(ruta_archivo)
    else:
        libro_excel = Workbook()

    # Verificar si se especificó una hoja nueva y válida
    if nombre_hoja_nueva:
        nombre_hoja = nombre_hoja_nueva
    else:
        # Si no se especifica, no se creará una nueva hoja
        print("No se especificó una hoja nueva. No se añadirá una hoja.")
        return

    # Crear o obtener la hoja
    nombre_hoja_final = agregar_hoja_excel(libro_excel, nombre_hoja)
    if not nombre_hoja_final:
        print("No se creará la hoja porque no se proporcionó un nombre válido.")
        return
    
    hoja = libro_excel[nombre_hoja_final]

    # Añadir cabecera solo si es una nueva hoja
    if hoja.max_row == 1:
        hoja.append(["ID", "Nombre de dato / archivo", "Formato", "Fecha de creación", "Fecha de modificación", "Ruta",
                     "Responsable del dato / archivo", "Propósito del dato / archivo",
                     "¿Quién tiene acceso al archivo / dato?", "Transferencia con 3ero / externos",
                     "Responsable de respaldo", "Fecha de respaldo", "Responsable de eliminación", "Fecha de eliminación"])

    # Iniciar cronómetro
    tiempo_inicio = time.time()

    # Guardar los datos en la hoja
    for id_archivo, archivo in enumerate(archivos, start=1):
        hoja.append([id_archivo] + list(archivo))

    # Ajustar el ancho de las columnas (esto es opcional)
    ajustar_ancho_columnas(hoja)

    try:
        # Guardar el archivo Excel
        libro_excel.save(ruta_archivo)
        tiempo_total = time.time() - tiempo_inicio
        hora_final = datetime.now()
        hora_formateada = hora_final.strftime("%H:%M:%S")

        print(f"Datos guardados exitosamente en el archivo Excel: {ruta_archivo}")
        print(f"Tiempo total para guardar en Excel: {tiempo_total:.2f} segundos a las {hora_formateada}")

        # Agregar el resumen del análisis a la lista global
        resumen_analisis_servidores.append({
            "Servidor IP": ip_servidor,
            "Ruta": ruta_servidor,
            "Archivo Excel": nombre_archivo,
            "Hoja": nombre_hoja_final,
            "Completado": True,
            "Hora Final": hora_formateada,
            "Duración": f"{tiempo_total:.2f} segundos"
        })

    except PermissionError as e:
        print(f"Error de permisos al guardar el archivo: {ruta_archivo} - {e}")
        # Agregar a la lista global que el análisis no se completó
        resumen_analisis_servidores.append({
            "Servidor IP": ip_servidor,
            "Ruta": ruta_servidor,
            "Archivo Excel": nombre_archivo,
            "Hoja": nombre_hoja_final,
            "Completado": False,
            "Hora Final": None,
            "Duración": None
        })

def mostrar_resumen_analisis():
    print("\nResumen del análisis de servidores:")
    for resumen in resumen_analisis_servidores:
        print(f"\nServidor IP: {resumen['Servidor IP']}")
        print(f"Ruta: {resumen['Ruta']}")
        print(f"Archivo Excel: {resumen['Archivo Excel']}")
        print(f"Hoja: {resumen['Hoja']}")
        print(f"Completado: {'Sí' if resumen['Completado'] else 'No'}")
        if resumen['Completado']:
            print(f"Hora de finalización: {resumen['Hora Final']}")
            print(f"Duración del análisis: {resumen['Duración']}")

def limpiar_nombre(nombre):
    caracteres_no_validos = ['\\', '/', '*', '[', ']', ':', '?']
    for caracter in caracteres_no_validos:
        nombre = nombre.replace(caracter, '_')
    return nombre[:30]

def establecer_conexion(host, username, use_private_key, private_key_path=None, password=None, port=22, passphrase=None):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    
    try:
        if use_private_key and private_key_path:
            pkey = paramiko.RSAKey.from_private_key_file(private_key_path, password=passphrase)
            ssh.connect(hostname=host, port=port, username=username, pkey=pkey)
        else:
            ssh.connect(hostname=host, port=port, username=username, password=password)
        print(f"Conexión establecida con {host}")
        return ssh
    except paramiko.AuthenticationException:
        print("Fallo en la autenticación")
    except paramiko.SSHException as sshException:
        print(f"Error al intentar conectar al host: {sshException}")
    except Exception as e:
        print(f"Error inesperado al intentar conectar: {e}")
        return None

# Función para verificar la hora límite
def verificar_hora_limite(hora_fin, hora_final_dt, resumen_operacion, nombre_archivo_excel, ip_servidor, ruta_servidor, detener_event):
    while not detener_event.is_set():
        ahora = datetime.now()

        # Comprobar si se ha alcanzado la hora límite (por defecto 5 AM)
        if ahora >= hora_fin:
            print("Se alcanzó la hora límite. Guardando el progreso y pausando la operación.")
            guardar_en_excel(resumen_operacion, nombre_archivo_excel, ip_servidor, ruta_servidor)
            mostrar_resumen_analisis()
            detener_event.set()  # Detener el proceso principal
            return
        
        # Comprobar si se ha alcanzado la hora final personalizada (si se proporcionó)
        if hora_final_dt and ahora >= hora_final_dt:
            print("Se alcanzó la hora final personalizada. Guardando el progreso y pausando la operación.")
            guardar_en_excel(resumen_operacion, nombre_archivo_excel, ip_servidor, ruta_servidor)
            mostrar_resumen_analisis()
            detener_event.set()  # Detener el proceso principal
            return
        
        # Esperar 1 segundo antes de volver a verificar
        time.sleep(1)

def iniciar_operacion_multiple(conexiones, nombre_archivo_excel,ip_servidor, ruta_servidor, hora_final=None):
    # Convertir hora_final en un objeto datetime solo si se proporciona
    if hora_final:
        hora_final_dt = datetime.combine(datetime.today(), hora_final)
    else:
        hora_final_dt = None

    # Obtener la hora actual y establecer hora_fin como 5 AM por defecto, si no se proporciona hora_final
    hora_inicio = datetime.now()
    hora_fin = hora_inicio.replace(hour=5, minute=0, second=0, microsecond=0)

    # Si la hora final proporcionada es anterior a la hora de inicio, asumir que corresponde al día siguiente
    if hora_final and hora_inicio.time() > hora_final:
        hora_final_dt += timedelta(days=1)
    
    # Si la hora de inicio es posterior a las 5 AM, asumir que la hora límite pertenece al día siguiente
    if hora_inicio.time() >= hora_fin.time():
        hora_fin += timedelta(days=1)

    # Lista para guardar el resumen de las operaciones
    resumen_operacion = []

    # Evento para detener el proceso cuando se alcance la hora límite
    detener_event = threading.Event()

    # Iniciar un hilo para verificar la hora límite cada minuto
    hilo_verificacion = threading.Thread(target=verificar_hora_limite, args=(hora_fin, hora_final_dt, resumen_operacion, nombre_archivo_excel, ip_servidor, ruta_servidor, detener_event))
    hilo_verificacion.start()

    # Comienza el ciclo de conexiones y análisis
    for conexion in conexiones:
        if detener_event.is_set():  # Detener el proceso si se alcanzó la hora límite
            break

        host, username, use_private_key, private_key_path, password, port, rutas, sistema_operativo, passphrase = conexion
        ssh_client = establecer_conexion(host, username, use_private_key, private_key_path, password, port, passphrase)
        if ssh_client:
            try:
                sftp_client = ssh_client.open_sftp()
                for carpeta, nombre_hoja_nueva, es_red_compartida in rutas:
                    if detener_event.is_set():  # Detener el proceso si se alcanzó la hora límite
                        break

                    # Operaciones de listar archivos según el sistema operativo
                    if es_red_compartida:
                        if sistema_operativo == "windows":
                            archivos = listar_archivos_remotos_windows_red_compartida(carpeta, detener_event)
                        else:
                            print("Las carpetas de red compartidas no son compatibles con sistemas operativos que no sean Windows.")
                            continue
                    else:
                        if sistema_operativo == "linux":
                            archivos = listar_archivos_remotos_linux(sftp_client, carpeta, detener_event)
                        elif sistema_operativo == "windows":
                            archivos = listar_archivos_remotos_windows(sftp_client, carpeta, detener_event)
                        else:
                            print("Sistema operativo no soportado.")
                            continue
                    
                    if archivos:
                        # Guardar datos de la operación en el resumen
                        resumen_operacion.append({
                            "host": host,
                            "carpeta": carpeta,
                            "archivos": archivos,
                            "hoja": nombre_hoja_nueva,
                            "hora_completado": datetime.now().strftime('%H:%M:%S')
                        })
                        guardar_en_excel(archivos, nombre_archivo_excel, host, carpeta, nombre_hoja_nueva)
                        print(f"Datos guardados exitosamente en la hoja '{nombre_hoja_nueva}' en {nombre_archivo_excel}.")
                    else:
                        print(f"No se encontraron archivos en la carpeta {carpeta}.")
            except Exception as e:
                print(f"Error al procesar la operación en {host}: {e}")
            finally:
                ssh_client.close()
        else:
            print(f"No se pudo establecer la conexión SSH con {host}.")
    
    # Asegurarse de que el hilo de verificación de hora se cierre correctamente
    detener_event.set()
    hilo_verificacion.join()

    # Al finalizar todas las conexiones, mostrar el resumen
    mostrar_resumen_analisis()

#Ejecucion hora_inicio
def esperar_hasta_hora_objetivo(hora_objetivo):
    ahora = datetime.now()
    objetivo = ahora.replace(hour=hora_objetivo.hour, minute=hora_objetivo.minute, second=0, microsecond=0)
    
    if ahora > objetivo:
        objetivo = objetivo + timedelta(days=1)
    
    tiempo_espera = (objetivo - ahora).total_seconds()
    print(f"Esperando {tiempo_espera / 60:.2f} minutos para ejecutar la tarea a las {objetivo.time()}.")
    time.sleep(tiempo_espera)

def main():
    def obtener_entrada_opcion(prompt, opciones_validas):
        """Obtiene una opción válida del usuario."""
        while True:
            entrada = input(prompt).strip().lower()
            if entrada in opciones_validas:
                return entrada
            else:
                print(f"Opción no válida. Por favor, ingrese {' o '.join(opciones_validas)}.")

    def obtener_entrada_numerica(prompt, valor_default):
        """Obtiene una entrada numérica válida del usuario."""
        while True:
            entrada = input(prompt).strip()
            if entrada == "":
                return valor_default
            if entrada.isdigit():
                return int(entrada)
            else:
                print("Entrada no válida. Por favor, ingrese un número.")

    def obtener_entrada_no_vacia(prompt):
        """Obtiene una entrada no vacía del usuario."""
        while True:
            entrada = input(prompt).strip()
            if entrada:
                return entrada
            else:
                print("La entrada no puede estar vacía.")

    def obtener_entrada_hora(prompt):
        """Obtiene una hora en formato HH:MM del usuario y valida el formato."""
        while True:
            entrada = input(prompt).strip()
            try:
                return datetime.strptime(entrada, "%H:%M").time()
            except ValueError:
                print("El formato de la hora no es válido. Asegúrese de usar HH:MM.")

    conexiones = []
    while True:
        host = obtener_entrada_no_vacia("Ingrese la dirección IP o hostname del servidor: ")
        username = obtener_entrada_no_vacia("Ingrese el nombre de usuario para la conexión SSH: ")
        use_private_key = obtener_entrada_opcion("¿Desea usar una clave privada para la autenticación? (s/n): ", ['s', 'n']) == 's'
        private_key_path = input("Ingrese la ruta al archivo de clave privada (deje vacío si no usa clave privada): ") or None
        passphrase = input("Ingrese la passphrase para la clave privada (deje vacío si no requiere passphrase): ") or None
        password = input("Ingrese la contraseña (en caso de necesitarla, en caso contrario dejar vacío): ") or None
        port = obtener_entrada_numerica("Ingrese el puerto del servidor SSH (por defecto 22): ", 22)

        while True:
            sistema_operativo = obtener_entrada_opcion("Ingrese el sistema operativo (linux/windows): ", ["linux", "windows"])
            if sistema_operativo in ["linux", "windows"]:
                break

        rutas = []
        nombre_archivo_excel = input("Ingrese el nombre del archivo Excel (deje vacío para 'Bitacora.xlsx'): ") or "Bitacora.xlsx"

        # Asegurar que el nombre del archivo tenga la extensión .xlsx
        if not nombre_archivo_excel.endswith(".xlsx"):
            nombre_archivo_excel += ".xlsx"

        while True:
            if os.path.exists(nombre_archivo_excel):
                try:
                    # Intentar cargar el archivo
                    workbook = openpyxl.load_workbook(nombre_archivo_excel)
                    print(f"Archivo '{nombre_archivo_excel}' cargado exitosamente.")
                    break
                except Exception as e:
                    # Si falla al cargar, preguntamos si desea crear uno nuevo o intentar otro
                    print(f"No se pudo cargar el archivo '{nombre_archivo_excel}'. Error: {e}")
                    respuesta = obtener_entrada_opcion("¿Desea crear uno nuevo? (s/n): ", ['s', 'n'])
                    if respuesta == 's':
                        workbook = openpyxl.Workbook()  # Crear un nuevo archivo Excel
                        workbook.save(nombre_archivo_excel)  # Guardar el nuevo archivo
                        print(f"Archivo '{nombre_archivo_excel}' creado exitosamente.")
                        break
                    else:
                        # Preguntar por un nuevo nombre de archivo
                        nombre_archivo_excel = input("Ingrese el nombre del archivo Excel (deje vacío para 'Bitacora.xlsx'): ") or "Bitacora.xlsx"
                        if not nombre_archivo_excel.endswith(".xlsx"):
                            nombre_archivo_excel += ".xlsx"
            else:
                # Si no existe, preguntar si desea crearlo
                print(f"El archivo '{nombre_archivo_excel}' no existe.")
                respuesta = obtener_entrada_opcion("¿Desea crear uno nuevo? (s/n): ", ['s', 'n'])
                if respuesta == 's':
                    workbook = openpyxl.Workbook()  # Crear un nuevo archivo Excel
                    workbook.save(nombre_archivo_excel)  # Guardar el nuevo archivo
                    print(f"Archivo '{nombre_archivo_excel}' creado exitosamente.")
                    break
                else:
                    # Preguntar por un nuevo nombre de archivo
                    nombre_archivo_excel = input("Ingrese el nombre del archivo Excel (deje vacío para 'Bitacora.xlsx'): ") or "Bitacora.xlsx"
                    if not nombre_archivo_excel.endswith(".xlsx"):
                        nombre_archivo_excel += ".xlsx"

        nombres_hojas = set(workbook.sheetnames)

        while True:
            carpeta = obtener_entrada_no_vacia("Ingrese la ruta remota a analizar: ")
            nombre_hoja_nueva = obtener_entrada_no_vacia("Ingrese el nombre de la hoja: ")

            if nombre_hoja_nueva in nombres_hojas:
                respuesta = obtener_entrada_opcion(f"La hoja '{nombre_hoja_nueva}' ya existe. ¿Desea continuar escribiendo ahí? (s/n): ", ['s', 'n'])
                if respuesta == 'n':
                    nombre_hoja_nueva = obtener_entrada_no_vacia("Ingrese un nuevo nombre para la hoja: ")
                    if nombre_hoja_nueva in nombres_hojas:
                        print(f"La hoja '{nombre_hoja_nueva}' también ya existe.")
                        continue
            else:
                print(f"Se creará una nueva hoja '{nombre_hoja_nueva}'.")

            nombres_hojas.add(nombre_hoja_nueva)
            es_red_compartida = obtener_entrada_opcion("¿Es esta una carpeta de red compartida? (s/n): ", ['s', 'n']) == 's'
            rutas.append((carpeta, nombre_hoja_nueva, es_red_compartida))
            otra_ruta = obtener_entrada_opcion("¿Desea agregar otra ruta a analizar para este servidor? (s/n): ", ['s', 'n'])
            if otra_ruta != 's':
                break
        
        conexiones.append((host, username, use_private_key, private_key_path, password, port, rutas, sistema_operativo, passphrase))
        otro_servidor = obtener_entrada_opcion("¿Desea agregar otro servidor a analizar? (s/n): ", ['s', 'n'])
        if otro_servidor != 's':
            break

    opcion = obtener_entrada_opcion("¿Desea ejecutar la operación ahora (1) o en una hora específica (2)? Ingrese 1 o 2: ", ['1', '2'])
    
    if opcion == '1':
        for conexion in conexiones:
            ip_servidor = conexion[0]  # El valor 'host' es ip_servidor
            rutas = conexion[6]  # Rutas es una lista de tuplas
            for ruta in rutas:
                ruta_servidor = ruta[0]  # El primer elemento de cada tupla en rutas es 'carpeta'
                iniciar_operacion_multiple([conexion], nombre_archivo_excel, ip_servidor, ruta_servidor, hora_final=None)

    elif opcion == '2':
        hora_inicio = obtener_entrada_hora("Ingrese la hora de inicio en formato HH:MM (por ejemplo, 22:00): ")
        print(f"Hora de inicio establecida a las {hora_inicio}.")
        
        hora_final_str = obtener_entrada_opcion("¿Desea agregar una hora final para limitar el tiempo de ejecución? (s/n): ", ['s', 'n'])
        if hora_final_str == 's':
            hora_final = obtener_entrada_hora("Ingrese la hora final en formato HH:MM (por ejemplo, 05:00): ")
            print(f"Hora final establecida a las {hora_final}.")
        else:
            hora_final = None

        esperar_hasta_hora_objetivo(hora_inicio)

        for conexion in conexiones:
            ip_servidor = conexion[0]  # El valor 'host' es ip_servidor
            rutas = conexion[6]  # Rutas es una lista de tuplas
            for ruta in rutas:
                ruta_servidor = ruta[0]  # El primer elemento de cada tupla en rutas es 'carpeta'
                iniciar_operacion_multiple([conexion], nombre_archivo_excel, ip_servidor, ruta_servidor, hora_final=hora_final)
    
    else:
        print("Opción no válida. Por favor, ingrese 1 o 2.")


if __name__ == "__main__":
    print(
'''
_____________________$$$
____________________$___$               
_____________________$$$
_____________________$_$
_____________________$_$
___________________$$$_$$$
_________________$$__$$$__$$$
_______________$$__$$$$$$$___$
______________$_______________$
_____________$_________________$
_____________$_________________$
_____________$_____$$$$$$$$$$$$$$$
_____________$____$_______________$
_____________$____$___$$$$$$$$$$$$$
_____________$___$___$___________$$$
_____________$___$___$_$$$___$$$__$$
_____________$___$___$_$$$___$$$__$$
_____________$___$___$___________$$$
_____________$____$___$$$$$$$$$$$$$
_____________$_____$$$$$$$$$$$$$$
_____________$_________________$
_____________$____$$$$$$$$$$$$$$
_____________$___$__$__$__$__$
_____________$__$$$$$$$$$$$$$$
_____________$__$___$__$__$__$
_____________$___$$$$$$$$$$$$$$$
____________$$$_________________$$$
__________$$___$$$_________$$$$$___$$
________$$________$$$$$$$$$__________$$$
_______$__$$_____________________$$$$___$$
____$$$$$___$$$$$$$$______$$$$$$$_______$_$
__$______$$_________$$$$$$______________$_$$
_$____$____$____________________________$_$_$
_$_____$___$______________$$$$$$$$$$$___$_$_$$
_$$$____$___$__$$$$$$$$$$$$__________$___$_$_$$
$___$$$$____$__$_____________________$___$_$$_$
$$$____$___$$__$_____________________$$__$_$__$
$___$__$__$$___$______________________$__$$$__$
$_____$$_$$____$_______________$$$____$__$_$__$

--- Proceso Automático Unificado para Listado e Inventariado de Nodos Activos ---
GIHUB: bender2997
''')
    main()
